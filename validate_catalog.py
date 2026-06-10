#!/usr/bin/env python3
"""Comprehensive validation of katalog.xlsx"""
import openpyxl
import re

KATALOG_PATH = '/home/doruk/Code/ekerler/katalog.xlsx'
FIYAT_DIR = '/home/doruk/Code/ekerler/fiyat'

wb = openpyxl.load_workbook(KATALOG_PATH)
ws = wb.active

errors = []
warnings = []

print(f"Toplam satır: {ws.max_row}")
print(f"Başlık hariç: {ws.max_row - 1}")
print()

# 1. Check all required columns
required_cols = {
    'stockCode': (0, 'A'),
    'label': (1, 'B'),
    'status': (2, 'C'),
    'mainCategory': (7, 'H'),
    'rootProductStockCode': (13, 'N'),
    'price1': (15, 'P'),
    'tax': (20, 'U'),
    'currencyAbbr': (21, 'V'),
    'stockAmount': (22, 'W'),
    'stockType': (23, 'X'),
}

print("=== 1. ZORUNLU SÜTUN KONTROLÜ ===")
for col_name, (idx, col_letter) in required_cols.items():
    empty = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx] is None or row[idx] == '':
            empty += 1
    if empty > 0:
        errors.append(f"Sütun {col_letter}({col_name}): {empty} boş satır")
        print(f"  ❌ {col_name}: {empty} boş")
    else:
        print(f"  ✅ {col_name}: Tam dolu")

# 2. Check KDV
print()
print("=== 2. KDV KONTROLÜ ===")
kdv_dist = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    kdv = row[20]
    kdv_dist[kdv] = kdv_dist.get(kdv, 0) + 1

for kdv, count in sorted(kdv_dist.items()):
    if kdv == 1:
        print(f"  ✅ KDV %1: {count} satır")
    else:
        errors.append(f"KDV %{kdv}: {count} satır (beklenen %1)")
        print(f"  ❌ KDV %{kdv}: {count} satır")

# 3. Check for duplicate stockCodes
print()
print("=== 3. TEKRARLAYAN STOCKCODE KONTROLÜ ===")
stockcodes = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    code = str(row[0]) if row[0] else ''
    if code:
        if code in stockcodes:
            stockcodes[code].append(row_idx)
        else:
            stockcodes[code] = [row_idx]

duplicates = {k: v for k, v in stockcodes.items() if len(v) > 1}
if duplicates:
    errors.append(f"{len(duplicates)} tekrarlayan stockCode bulundu")
    for code, rows in list(duplicates.items())[:5]:
        print(f"  ❌ {code}: satırlar {rows}")
else:
    print("  ✅ Tekrarlayan stockCode yok")

# 4. Check variant structure
print()
print("=== 4. VARYANT YAPISI KONTROLÜ ===")
parent_rows = {}
variant_rows = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    code = str(row[0]) if row[0] else ''
    root = row[13]
    label = str(row[1]) if row[1] else ''
    
    if root == 0 or root is None or root == '':
        parent_rows[code] = {'row': row_idx, 'label': label, 'variants': []}
    else:
        variant_rows[code] = {'row': row_idx, 'root': str(root), 'label': label}

orphan_variants = 0
for code, info in variant_rows.items():
    if info['root'] not in parent_rows:
        orphan_variants += 1
        if orphan_variants <= 3:
            errors.append(f"Orphan varyant: {code} (root={info['root']})")
            print(f"  ❌ Orphan varyant: {code} (root={info['root']}, satır={info['row']})")

if orphan_variants == 0:
    print("  ✅ Tüm varyantların parent'i var")
else:
    print(f"  ❌ Toplam {orphan_variants} orphan varyant")

# Check variant consistency
inconsistent = 0
for code, info in parent_rows.items():
    parent_price = None
    for v_code, v_info in variant_rows.items():
        if v_info['root'] == code:
            # Get variant price
            v_price = ws.cell(row=v_info['row'], column=16).value
            v_label = str(ws.cell(row=v_info['row'], column=2).value)
            
            if parent_price is None:
                parent_price = ws.cell(row=info['row'], column=16).value
            
            # Check if variant price is proportional
            if v_label.isdigit() and parent_price:
                expected = round(parent_price * int(v_label) / 1000, 2)
                if abs(v_price - expected) > 0.5:
                    inconsistent += 1
                    if inconsistent <= 3:
                        errors.append(f"Fiyat tutarsız: {v_code} (beklenen {expected}, gerçek {v_price})")

if inconsistent == 0:
    print("  ✅ Tüm varyant fiyatları tutarlı")
else:
    print(f"  ❌ {inconsistent} tutarsız varyant fiyatı")

# 5. Check SEO descriptions
print()
print("=== 5. SEO AÇIKLAMA KONTROLÜ ===")
empty_seo = 0
short_seo = 0
for code, info in parent_rows.items():
    row_idx = info['row']
    details = ws.cell(row=row_idx, column=31).value
    if not details or len(str(details)) < 10:
        empty_seo += 1
        if empty_seo <= 3:
            errors.append(f"SEO eksik: satır {row_idx} ({info['label']})")
    elif len(str(details)) < 50:
        short_seo += 1

print(f"  Boş/çok kısa SEO: {empty_seo} ürün")
print(f"  Kısa SEO (<50 karakter): {short_seo} ürün")
if empty_seo == 0:
    print("  ✅ Tüm parent ürünlerde SEO var")

# 6. Check brands
print()
print("=== 6. MARKA KONTROLÜ ===")
brands = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    b = row[3]
    if b:
        brands[str(b)] = brands.get(str(b), 0) + 1

empty_brand = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if not row[3])
print(f"  Boş marka: {empty_brand} satır")
print(f"  Farklı marka: {len(brands)}")

# Check for suspicious brands
suspicious = ['Siyah', 'Beyaz', 'Kirmizi', 'Yesil', 'Sari', 'Meyve', 'Kuru', 
              'Cilek', 'Erik', 'Dut', 'Elma', 'Mango', 'Kivi', 'Visne',
              'Seftali', 'Igde', 'Incir', 'Malatya', 'Gun', 'Dal',
              'Sukkari', 'Cennet', 'Kus', 'Extra', 'Karcı']

found_suspicious = []
for brand in suspicious:
    if brand in brands:
        found_suspicious.append(brand)

if found_suspicious:
    errors.append(f"Şüpheli markalar: {found_suspicious}")
    print(f"  ❌ Şüpheli markalar: {found_suspicious}")
else:
    print("  ✅ Şüpheli marka yok")

# 7. Check file coverage
print()
print("=== 7. DOSYA KAPSAM KONTROLÜ ===")
import os
files = {
    'baharat tam liste.xlsx': 113,
    'KURUYEMIS.xlsx': 101,
    'ŞARKÜTER.xlsx': 83,
    'BAKLIYAT.xlsx': 30,
    'HELVA-TAHİN-PEKMEZ.xlsx': 13,
    'meyve kurusu.xlsx': 57,
    'ŞEKERLEME.xlsx': 15,
}

total_expected = sum(files.values())
print(f"  Beklenen toplam ürün: {total_expected}")
print(f"  Katalogdaki parent satır: {len(parent_rows)}")
print(f"  Orijinal ürünler (satır 2-59): ~58")

if len(parent_rows) >= total_expected + 50:  # +50 for original products
    print("  ✅ Tüm dosyalar işlenmiş görünüyor")
else:
    print(f"  ⚠️ Beklenenden az ürün var")

# Summary
print()
print("=" * 60)
print("ÖZET")
print("=" * 60)
if errors:
    print(f"❌ {len(errors)} KRİTİK HATA BULUNDU:")
    for i, err in enumerate(errors[:10], 1):
        print(f"  {i}. {err}")
    if len(errors) > 10:
        print(f"  ... ve {len(errors)-10} hata daha")
else:
    print("✅ KRİTİK HATA YOK!")

if warnings:
    print(f"\n⚠️ {len(warnings)} uyarı var")
else:
    print("\n✅ Uyarı yok")

print(f"\nToplam: {len(errors)} hata, {len(warnings)} uyarı")
